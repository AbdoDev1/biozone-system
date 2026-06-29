from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from products.models import Product, ProductUnit, Category
from products.forms import ProductForm, ProductUnitForm
from inventory.models import Inventory, StockMovement
import openpyxl


def staff_required(view_func):
    def wrapper(request, *args, **kwargs):
        if not request.user.is_authenticated or request.user.role not in ['ADMIN', 'WAREHOUSE']:
            return redirect('staff:login')
        return view_func(request, *args, **kwargs)
    return wrapper


@staff_required
def product_list(request):
    products = Product.objects.select_related('category').prefetch_related('units').all()
    categories = Category.objects.filter(is_active=True)
    selected_category = request.GET.get('category', '')
    search_q = request.GET.get('q', '')
    if selected_category:
        products = products.filter(category__slug=selected_category)
    if search_q:
        products = products.filter(name_ar__icontains=search_q)
    return render(request, 'staff/products/list.html', {
        'products': products,
        'categories': categories,
        'selected_category': selected_category,
        'search_q': search_q,
    })


@staff_required
def product_add(request):
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES)
        unit_form = ProductUnitForm(request.POST)
        if form.is_valid() and unit_form.is_valid():
            product = form.save()
            unit = unit_form.save(commit=False)
            unit.product = product
            unit.save()
            initial_stock = unit_form.cleaned_data.get('initial_stock') or 0
            inventory = Inventory.objects.create(
                product_unit=unit,
                quantity=initial_stock,
                reserved=0,
                min_quantity=0,
            )
            if initial_stock > 0:
                StockMovement.objects.create(
                    inventory=inventory,
                    movement_type='IN',
                    quantity=initial_stock,
                    note='كمية ابتدائية عند إضافة المنتج',
                    created_by=request.user,
                )
            messages.success(request, f'تم إضافة المنتج "{product.name_ar}" بنجاح.')
            return redirect('staff:product_list')
    else:
        form = ProductForm()
        unit_form = ProductUnitForm()
    return render(request, 'staff/products/form.html', {
        'form': form,
        'unit_form': unit_form,
        'title': 'إضافة منتج جديد',
        'is_edit': False,
    })


@staff_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    unit = product.units.first()
    if request.method == 'POST':
        form = ProductForm(request.POST, request.FILES, instance=product)
        unit_form = ProductUnitForm(request.POST, instance=unit) if unit else ProductUnitForm(request.POST)
        if form.is_valid() and unit_form.is_valid():
            form.save()
            edited_unit = unit_form.save(commit=False)
            edited_unit.product = product
            edited_unit.save()
            messages.success(request, f'تم تعديل المنتج "{product.name_ar}" بنجاح.')
            return redirect('staff:product_list')
    else:
        form = ProductForm(instance=product)
        unit_form = ProductUnitForm(instance=unit) if unit else ProductUnitForm()
    return render(request, 'staff/products/form.html', {
        'form': form,
        'unit_form': unit_form,
        'title': f'تعديل: {product.name_ar}',
        'is_edit': True,
        'product': product,
    })


@staff_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)

    has_stock = any(
        unit.inventory.quantity > 0
        for unit in product.units.all()
        if hasattr(unit, 'inventory')
    )

    if request.method == 'POST':
        name = product.name_ar
        if has_stock:
            product.is_active = False
            product.save()
            messages.warning(request, f'المنتج "{name}" له مخزون — تم تعطيله بدل الحذف.')
        else:
            product.delete()
            messages.success(request, f'تم حذف المنتج "{name}".')
        return redirect('staff:product_list')

    return render(request, 'staff/products/delete.html', {
        'product': product,
        'has_stock': has_stock,
    })


@staff_required
def import_products(request):
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        if not excel_file:
            messages.error(request, 'اختار ملف Excel أولاً.')
            return redirect('staff:import_products')
        if not excel_file.name.endswith('.xlsx'):
            messages.error(request, 'الملف لازم يكون .xlsx')
            return redirect('staff:import_products')
        try:
            wb = openpyxl.load_workbook(excel_file)
            ws = wb.active
            headers = [str(cell.value).strip() if cell.value else '' for cell in ws[1]]
            required_headers = ['name_ar', 'category_slug', 'unit_name', 'unit_price', 'initial_stock']
            missing = [h for h in required_headers if h not in headers]
            if missing:
                messages.error(request, f'الأعمدة دي ناقصة: {", ".join(missing)}')
                return redirect('staff:import_products')
            idx = {h: headers.index(h) for h in headers if h}
            success_count = 0
            error_rows = []
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                try:
                    name_ar = str(row[idx['name_ar']]).strip() if row[idx['name_ar']] else ''
                    category_slug = str(row[idx['category_slug']]).strip() if row[idx['category_slug']] else ''
                    unit_name = str(row[idx['unit_name']]).strip() if row[idx['unit_name']] else ''
                    unit_price = float(row[idx['unit_price']]) if row[idx['unit_price']] else 0
                    initial_stock = int(row[idx['initial_stock']]) if row[idx['initial_stock']] else 0
                    name_en = str(row[idx['name_en']]).strip() if 'name_en' in idx and row[idx['name_en']] else ''
                    manufacturer = str(row[idx['manufacturer']]).strip() if 'manufacturer' in idx and row[idx['manufacturer']] else ''
                    wholesale_price = float(row[idx['wholesale_price']]) if 'wholesale_price' in idx and row[idx['wholesale_price']] else None
                    wholesale_min_qty = int(row[idx['wholesale_min_qty']]) if 'wholesale_min_qty' in idx and row[idx['wholesale_min_qty']] else 0
                    if not name_ar or not category_slug or not unit_name or not unit_price:
                        error_rows.append(f'سطر {row_num}: بيانات ناقصة')
                        continue
                    try:
                        category = Category.objects.get(slug=category_slug)
                    except Category.DoesNotExist:
                        error_rows.append(f'سطر {row_num}: القسم "{category_slug}" مش موجود')
                        continue
                    product, created = Product.objects.get_or_create(
                        name_ar=name_ar,
                        defaults={
                            'category': category,
                            'name_en': name_en,
                            'manufacturer': manufacturer,
                            'is_active': True,
                        }
                    )
                    if not created:
                        product.category = category
                        product.manufacturer = manufacturer
                        product.save()
                    unit, _ = ProductUnit.objects.update_or_create(
                        product=product,
                        size='S',
                        defaults={
                            'name': unit_name,
                            'unit_price': unit_price,
                            'wholesale_price': wholesale_price,
                            'wholesale_min_qty': wholesale_min_qty,
                            'wholesale_mode': 'FULL',
                            'qty_in_small': 1,
                        }
                    )
                    inventory, _ = Inventory.objects.get_or_create(
                        product_unit=unit,
                        defaults={'quantity': 0, 'reserved': 0, 'min_quantity': 0}
                    )
                    if initial_stock > 0:
                        inventory.quantity += initial_stock
                        inventory.save()
                        StockMovement.objects.create(
                            inventory=inventory,
                            movement_type='IN',
                            quantity=initial_stock,
                            note='إضافة من ملف Excel',
                            created_by=request.user,
                        )
                    success_count += 1
                except Exception as e:
                    error_rows.append(f'سطر {row_num}: خطأ — {str(e)}')
                    continue
            if success_count:
                messages.success(request, f'تم إضافة/تحديث {success_count} منتج بنجاح.')
            for err in error_rows:
                messages.warning(request, err)
        except Exception as e:
            messages.error(request, f'خطأ في قراءة الملف: {str(e)}')
    return render(request, 'staff/products/import.html')


@staff_required
def download_template(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'المنتجات'
    headers = ['name_ar', 'name_en', 'category_slug', 'manufacturer',
               'unit_name', 'unit_price', 'wholesale_price', 'wholesale_min_qty', 'initial_stock']
    ws.append(headers)
    ws.append(['قفازات لاتكس', 'Latex Gloves', 'gloves', 'Medline', 'علبة', 25.00, 20.00, 10, 100])
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20
    from django.http import HttpResponse
    import io
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="biozone_products_template.xlsx"'
    return response
